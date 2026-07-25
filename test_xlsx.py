import os
import csv
import time
import shutil

from lib import g
from lib import io
from lib import llm

from openpyxl import load_workbook

def xlsx_to_csv():
    xlsx = "/home/ubuntu/vault/terrawhisper/data/suppliers/INTEGRITY_Export_20260701.xlsx"
    csv_file = "/home/ubuntu/vault/terrawhisper/data/suppliers/INTEGRITY_Export_20260701.csv"
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    with open(csv_file, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    wb.close()


def parse_csv():
    filepath = "/home/ubuntu/vault/terrawhisper/data/suppliers/INTEGRITY_Export_20260701.csv"
    with open(filepath, encoding="utf-8", newline="") as f:
        for row in csv.reader(f):
            print(row, flush=True)
            quit()

def run():
    print('SUPPLIER >> PARSE >> usda_organic')

    # xlsx_to_csv() ### ONE TIME ONLY
    parse_csv()

