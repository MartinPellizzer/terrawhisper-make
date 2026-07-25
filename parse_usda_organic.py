import os
import re
import csv
import time
import json
import shutil
import sqlite3
import unicodedata

from lib import g
from lib import io
from lib import llm

from openpyxl import load_workbook

import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse

import resolve_utils
import derive_utils

def xlsx_to_csv():
    xlsx = "/home/ubuntu/vault/terrawhisper/data/organizations/fetch/usda_organic/INTEGRITY_Export_20260701.xlsx"
    csv_file = "/home/ubuntu/vault/terrawhisper/data/organizations/fetch/usda_organic/INTEGRITY_Export_20260701.csv"
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    with open(csv_file, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    wb.close()

def normalize_products_names(text):
    if not text:
        return []
    # Normalize Unicode and lowercase
    text = unicodedata.normalize("NFKC", text).lower()
    # Remove "other:" prefix
    text = re.sub(r"\bother\s*:\s*", "", text)
    # Split on commas, periods, newlines, and semicolons
    products = re.split(r"[,.;\n]+", text)
    # Normalize whitespace and remove non-alphanumeric characters
    products = [
        re.sub(r"[^a-z0-9\s-]", "", product).strip()
        for product in products
    ]
    # Remove empty entries and duplicates
    return list(dict.fromkeys(
        product for product in products if product
    ))

def scrape_homepage(url):
    if not url:
        return None
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    try:
        response = requests.get(
            url,
            timeout=15,
            headers={
                "User-Agent": "TerraWhisperBot/1.0 (+https://terrawhisper.com/bot)"
            }
        )
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        # Remove non-content elements
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        text = " ".join(soup.stripped_strings)
        return {
            "url": response.url,
            "title": soup.title.get_text(strip=True) if soup.title else "",
            "text": text,
        }
    except requests.RequestException:
        return None

def parse_csv():
    plants_popular = derive_utils.plants_popular_get()[:1]

    '''
    filepath = "/home/ubuntu/vault/terrawhisper/data/organizations/fetch/usda_organic/INTEGRITY_Export_20260701.csv"
    items = io.csv_to_dict(filepath, delimiter=',')
    found = []
    for item_i, item in enumerate(items[2:]):
        print(f'{item_i}/{len(items)}')
        website_url = item.get("op_url")
        if website_url:
            print(website_url)
            found.append(website_url)
            website_content = scrape_homepage(website_url)
            print(website_content)
            quit()
    '''
    ###
    filepath = "/home/ubuntu/vault/terrawhisper/data/organizations/fetch/usda_organic/INTEGRITY_Export_20260701.csv"
    items = io.csv_to_dict(filepath, delimiter=',')
    wcvp_folderpath = f'{g.VAULT_FOLDERPATH}/terrawhisper/data/reference/wcvp/wcvp.db'
    wcvp_conn = sqlite3.connect(wcvp_folderpath)
    found_businesses = []
    for item_i, item in enumerate(items[2:]):
        print(f'{item_i}/{len(items)}')
        text = " ".join([
            item.get("CR_CertifiedProducts", ""),
            item.get("LS_CertifiedProducts", ""),
            item.get("WC_CertifiedProducts", ""),
            item.get("Han_CertifiedProducts", ""),
        ]).lower()
        website_url = item.get("op_url")
        print(website_url)
        # quit()
        # print(text)
        names = normalize_products_names(text)
        for name in names:
            found = False
            for plant in plants_popular:
                if plant['plant_canonical_name'].lower().strip() == name.strip().lower():
                    found = True
                    break
            if found:
                print(name)
                # print(resolved)
                print()
                # quit()
                found_businesses.append({
                    'name': name,
                    # 'resolved': resolved,
                    'organization': item,
                })
                break
            
            # resolved = resolve_utils.resolve_plant_accepted(wcvp_conn, name)
            '''
            if resolved != None:
                print(name)
                print(resolved)
                print()
                # quit()
                found.append({
                    'name': name,
                    'resolved': resolved,
                    'organization': item,
                })
                break
            '''
        if len(found_businesses) >= 3:
            break
        # print(item, flush=True)
        # quit()
    for item in found_businesses:
        print(json.dumps(item, indent=4))
    wcvp_conn.close()

def run():
    print('ORGANIZATIONS >> PARSE >> usda_organic')

    # xlsx_to_csv() ### ONE TIME ONLY
    parse_csv()

